"""
extract_all.py
--------------
Extraction générique de TOUS les tableaux de TOUS les PDFs
pour TOUS les dossiers sociétés dans /financial_scraper_app/input/

Structure attendue :
    input/
        GROUPE_POULINA/
            POULINA_2014.pdf
            POULINA_2015.pdf
            ...
        AUTRE_SOCIETE/
            AUTRE_2020.pdf
            ...

Résultat produit :
    output/
        GROUPE_POULINA/
            POULINA_tables.xlsx   ← 1 onglet par fichier+tableau  (ex: 2024_P02_T00)
            POULINA_tables.pkl    ← cache pickle pour mode incrémental
        AUTRE_SOCIETE/
            AUTRE_tables.xlsx
            AUTRE_tables.pkl

Dépendances :
    pip install pdfplumber openpyxl

Usage :
    # Tout extraire
    python extract_all.py

    # Un seul dossier société
    python extract_all.py --company GROUPE_POULINA

    # Une seule année
    python extract_all.py --year 2024

    # Chemins personnalisés
    python extract_all.py --input /mon/chemin/input --output /mon/chemin/output

    # Forcer la ré-extraction même si déjà en cache
    python extract_all.py --force
"""

import argparse
import pickle
import re
import sys
from pathlib import Path

import openpyxl
import pdfplumber
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# ─────────────────────────────────────────────────────────────────────────────
# Styles Excel
# ─────────────────────────────────────────────────────────────────────────────

_TITLE_FILL  = PatternFill("solid", start_color="1F4E79")
_HEADER_FILL = PatternFill("solid", start_color="2E75B6")
_ALT_FILL    = PatternFill("solid", start_color="EBF1DE")
_thin        = Side(style="thin", color="BFBFBF")
_BORDER      = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


# ─────────────────────────────────────────────────────────────────────────────
# Utilitaires
# ─────────────────────────────────────────────────────────────────────────────

def clean(val: object) -> str:
    """Nettoie une valeur brute extraite du PDF."""
    if val is None:
        return ""
    return str(val).strip().replace("\n", " ")


def is_number_str(val: str) -> bool:
    """Renvoie True si val est une chaîne représentant un nombre (espaces tolérés)."""
    v = val.replace(" ", "").replace("\xa0", "")
    return bool(v.lstrip("-").replace(".", "").isdigit()) and len(v.lstrip("-")) > 0


def to_number(val: str) -> int | float | str:
    """Tente de convertir une chaîne en int ou float."""
    v = val.replace(" ", "").replace("\xa0", "")
    try:
        return int(v)
    except ValueError:
        try:
            return float(v)
        except ValueError:
            return val


def infer_year_from_filename(pdf_path: Path) -> int | None:
    """Extrait l'année depuis le nom du fichier (ex: POULINA_2024.pdf → 2024)."""
    match = re.search(r"(\d{4})", pdf_path.stem)
    return int(match.group(1)) if match else None


def extract_prefix_from_filename(pdf_path: Path) -> str:
    """Extrait le préfixe sans l'année (ex: POULINA_2024 → POULINA)."""
    return re.sub(r"[_\-]?\d{4}$", "", pdf_path.stem)


# ─────────────────────────────────────────────────────────────────────────────
# Extraction PDF → liste de DataFrames (sous forme de listes de listes)
# ─────────────────────────────────────────────────────────────────────────────

def extract_tables_from_pdf(
    pdf_path: Path,
    min_rows: int = 2,
    min_cols: int = 2,
) -> list[dict]:
    """
    Extrait tous les tableaux d'un PDF avec pdfplumber.

    Retourne une liste de dicts :
        {
            "page":    numéro de page (1-based),
            "table":   index du tableau dans la page (0-based),
            "n_rows":  nombre de lignes,
            "n_cols":  nombre de colonnes,
            "data":    list[list[str]]  — données nettoyées,
        }
    """
    results = []

    try:
        with pdfplumber.open(pdf_path) as pdf:
            for page_num, page in enumerate(pdf.pages, 1):
                tables = page.extract_tables()
                for tbl_idx, table in enumerate(tables):
                    if len(table) < min_rows:
                        continue
                    # Vérifier largeur minimale
                    max_cols = max(len(row) for row in table) if table else 0
                    if max_cols < min_cols:
                        continue

                    # Nettoyer toutes les cellules
                    cleaned = [
                        [clean(cell) for cell in row]
                        for row in table
                    ]

                    results.append({
                        "page":   page_num,
                        "table":  tbl_idx,
                        "n_rows": len(cleaned),
                        "n_cols": max_cols,
                        "data":   cleaned,
                    })

    except Exception as exc:
        print(f"    ❌ Erreur lecture {pdf_path.name} : {exc}")

    return results


# ─────────────────────────────────────────────────────────────────────────────
# Écriture Excel
# ─────────────────────────────────────────────────────────────────────────────

def _sheet_name(year: int, page: int, tbl_idx: int) -> str:
    """Génère un nom d'onglet unique et valide Excel (≤31 chars)."""
    return f"{year}_P{page:02d}_T{tbl_idx:02d}"[:31]


def write_table_to_sheet(ws, tbl_info: dict, year: int) -> None:
    """Écrit un tableau brut dans un onglet Excel avec mise en forme minimale."""
    data   = tbl_info["data"]
    page   = tbl_info["page"]
    t_idx  = tbl_info["table"]
    n_cols = tbl_info["n_cols"]

    # ── Ligne titre ──────────────────────────────────────────────────
    from openpyxl.utils import get_column_letter
    last_col = get_column_letter(max(n_cols, 1))
    ws.merge_cells(f"A1:{last_col}1")
    c = ws["A1"]
    c.value = f"Année {year}  |  Page {page}  |  Tableau {t_idx}"
    c.font  = Font(bold=True, color="FFFFFF", size=11)
    c.fill  = _TITLE_FILL
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 20

    # ── Données ──────────────────────────────────────────────────────
    for ri, row in enumerate(data, 2):
        is_first = (ri == 2)   # Première ligne → traiter comme en-tête potentiel

        # Padding si la ligne est plus courte que n_cols
        padded = row + [""] * (n_cols - len(row))

        for ci, raw in enumerate(padded, 1):
            val = to_number(raw) if is_number_str(raw) else raw
            c   = ws.cell(row=ri, column=ci, value=val)
            c.border = _BORDER

            if is_first:
                c.font  = Font(bold=True, color="FFFFFF", size=10)
                c.fill  = _HEADER_FILL
                c.alignment = Alignment(horizontal="center", wrap_text=True)
            else:
                c.font = Font(size=9)
                # Alternance de couleur sur les lignes paires
                if ri % 2 == 0:
                    c.fill = _ALT_FILL
                if ci > 1 and isinstance(val, (int, float)):
                    c.number_format = "#,##0"
                    c.alignment = Alignment(horizontal="right")
                elif ci == 1:
                    c.alignment = Alignment(wrap_text=True)

        ws.row_dimensions[ri].height = 15

    # ── Largeurs de colonnes ─────────────────────────────────────────
    ws.column_dimensions["A"].width = 40
    for ci in range(2, n_cols + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 16

    ws.freeze_panes = "A3"


def save_excel(
    all_data: dict[int, list[dict]],
    output_path: Path,
    prefix: str,
) -> None:
    """
    Écrit all_data dans un fichier Excel.
    all_data : {année: [tbl_info, ...]}
    """
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ── Onglet index ─────────────────────────────────────────────────
    ws_idx = wb.create_sheet("📋 Index")
    _build_index_sheet(ws_idx, all_data, prefix)

    # ── Un onglet par tableau ─────────────────────────────────────────
    total = 0
    for year in sorted(all_data.keys()):
        for tbl_info in all_data[year]:
            sname = _sheet_name(year, tbl_info["page"], tbl_info["table"])
            ws    = wb.create_sheet(sname)
            write_table_to_sheet(ws, tbl_info, year)
            total += 1

    wb.save(output_path)
    print(f"  📊 Excel → {output_path.name}  ({total} tableaux, {len(all_data)} années)")


def _build_index_sheet(ws, all_data: dict, prefix: str) -> None:
    """Onglet récapitulatif avec lien vers chaque tableau."""
    from openpyxl.utils import get_column_letter

    headers = ["Onglet", "Année", "Page PDF", "N° Tableau", "Lignes", "Colonnes", "Aperçu (1ère ligne)"]
    ws.merge_cells(f"A1:{get_column_letter(len(headers))}1")
    c = ws["A1"]
    c.value = f"INDEX — {prefix}"
    c.font  = Font(bold=True, color="FFFFFF", size=12)
    c.fill  = _TITLE_FILL
    c.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 22

    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=ci, value=h)
        c.font  = Font(bold=True, color="FFFFFF")
        c.fill  = _HEADER_FILL
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.row_dimensions[2].height = 28

    ri = 3
    for year in sorted(all_data.keys()):
        for tbl_info in all_data[year]:
            sname   = _sheet_name(year, tbl_info["page"], tbl_info["table"])
            preview = " | ".join(
                v for v in (tbl_info["data"][0] if tbl_info["data"] else []) if v
            )[:80]
            row_vals = [
                sname, year, tbl_info["page"], tbl_info["table"],
                tbl_info["n_rows"], tbl_info["n_cols"], preview,
            ]
            fill = _ALT_FILL if ri % 2 == 0 else PatternFill("solid", start_color="FFFFFF")
            for ci, val in enumerate(row_vals, 1):
                c = ws.cell(row=ri, column=ci, value=val)
                c.font  = Font(size=9)
                c.fill  = fill
                c.alignment = Alignment(horizontal="center" if ci != 7 else "left")
            ri += 1

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 8
    ws.column_dimensions["F"].width = 10
    ws.column_dimensions["G"].width = 60
    ws.freeze_panes = "A3"


# ─────────────────────────────────────────────────────────────────────────────
# Cache pickle (mode incrémental)
# ─────────────────────────────────────────────────────────────────────────────

def load_pkl(pkl_path: Path) -> dict:
    if pkl_path.exists():
        with open(pkl_path, "rb") as f:
            return pickle.load(f)
    return {}


def save_pkl(data: dict, pkl_path: Path) -> None:
    pkl_path.parent.mkdir(parents=True, exist_ok=True)
    with open(pkl_path, "wb") as f:
        pickle.dump(data, f)
    print(f"  🥒 PKL   → {pkl_path.name}")


# ─────────────────────────────────────────────────────────────────────────────
# Traitement d'un dossier société
# ─────────────────────────────────────────────────────────────────────────────

def process_company_folder(
    company_dir: Path,
    output_dir: Path,
    year_filter: int | None = None,
    force: bool = False,
) -> None:
    """
    Traite tous les PDFs d'un dossier société et produit Excel + PKL.
    """
    # Découvrir les PDFs
    pdfs = sorted(company_dir.glob("*.pdf"))
    if not pdfs:
        print(f"  ⚠️  Aucun PDF dans {company_dir}")
        return

    # Filtrer par année si demandé
    if year_filter:
        pdfs = [p for p in pdfs if str(year_filter) in p.stem]

    if not pdfs:
        print(f"  ⚠️  Aucun PDF pour l'année {year_filter}")
        return

    # Déduire préfixe (ex: POULINA)
    prefix    = extract_prefix_from_filename(pdfs[0])
    out_dir   = output_dir / company_dir.name
    pkl_path  = out_dir / f"{prefix}_tables.pkl"
    xlsx_path = out_dir / f"{prefix}_tables.xlsx"

    # Charger le cache existant
    all_data: dict[int, list[dict]] = {} if force else load_pkl(pkl_path)

    # ── Traiter chaque PDF ────────────────────────────────────────────
    for pdf_path in pdfs:
        year = infer_year_from_filename(pdf_path)
        if year is None:
            print(f"    ⚠️  Impossible de déduire l'année de : {pdf_path.name}, ignoré.")
            continue

        # Mode incrémental : sauter si déjà traité
        if year in all_data and not force:
            print(f"  ⏭️  {pdf_path.name} déjà en cache ({len(all_data[year])} tableaux)")
            continue

        print(f"  📄 {pdf_path.name} … ", end="", flush=True)
        tables = extract_tables_from_pdf(pdf_path)
        print(f"{len(tables)} tableau(x)")

        all_data[year] = tables

    if not all_data:
        print("  ⚠️  Aucune donnée extraite.")
        return

    # ── Sauvegarder ──────────────────────────────────────────────────
    save_excel(all_data, xlsx_path, prefix)
    save_pkl(all_data, pkl_path)


# ─────────────────────────────────────────────────────────────────────────────
# Point d'entrée principal
# ─────────────────────────────────────────────────────────────────────────────

def run(
    input_dir: Path,
    output_dir: Path,
    company_filter: str | None = None,
    year_filter: int | None = None,
    force: bool = False,
) -> None:
    """
    Parcourt input_dir, traite chaque sous-dossier société.
    """
    print(f"\n{'='*60}")
    print(f"📂 Input  : {input_dir}")
    print(f"📤 Output : {output_dir}")
    if company_filter:
        print(f"🔍 Filtre société : {company_filter}")
    if year_filter:
        print(f"🔍 Filtre année   : {year_filter}")
    print(f"{'='*60}\n")

    # Lister les dossiers sociétés
    company_dirs = [d for d in sorted(input_dir.iterdir()) if d.is_dir()]

    if not company_dirs:
        print(f"❌ Aucun sous-dossier trouvé dans {input_dir}")
        sys.exit(1)

    # Filtrer par société
    if company_filter:
        company_dirs = [
            d for d in company_dirs
            if company_filter.upper() in d.name.upper()
        ]
        if not company_dirs:
            print(f"❌ Aucun dossier correspondant à '{company_filter}'")
            sys.exit(1)

    print(f"🏢 {len(company_dirs)} société(s) à traiter\n")

    for company_dir in company_dirs:
        print(f"\n{'─'*60}")
        print(f"🏢  {company_dir.name}")
        print(f"{'─'*60}")
        process_company_folder(
            company_dir=company_dir,
            output_dir=output_dir,
            year_filter=year_filter,
            force=force,
        )

    print(f"\n{'='*60}")
    print("✅ Extraction terminée.")
    print(f"{'='*60}\n")


# ─────────────────────────────────────────────────────────────────────────────
# CLI
# ─────────────────────────────────────────────────────────────────────────────

# Racine du projet = 2 niveaux au-dessus de ce script
# financial_scraper_app/
# └── app/
#     └── data_Extraction/
#         └── Extract_poulina.py   ← ici
_SCRIPT_DIR  = Path(__file__).resolve().parent          # data_Extraction/
_PROJECT_ROOT = _SCRIPT_DIR.parents[1]                  # financial_scraper_app/
_DEFAULT_INPUT  = _PROJECT_ROOT / "input"
_DEFAULT_OUTPUT = _PROJECT_ROOT / "output"


def _parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Extraction générique PDF → Excel pour tous les dossiers sociétés.",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
    )
    parser.add_argument(
        "--input",
        default=str(_DEFAULT_INPUT),
        help="Dossier racine contenant les sous-dossiers sociétés.",
    )
    parser.add_argument(
        "--output",
        default=str(_DEFAULT_OUTPUT),
        help="Dossier de sortie (créé si inexistant).",
    )
    parser.add_argument(
        "--company",
        default=None,
        help="Traiter uniquement ce dossier société (recherche partielle, ex: POULINA).",
    )
    parser.add_argument(
        "--year",
        type=int,
        default=None,
        help="Traiter uniquement cette année (ex: 2024).",
    )
    parser.add_argument(
        "--force",
        action="store_true",
        help="Forcer la ré-extraction même si le cache PKL existe.",
    )
    return parser.parse_args()


if __name__ == "__main__":
    args = _parse_args()
    run(
        input_dir=Path(args.input),
        output_dir=Path(args.output),
        company_filter=args.company,
        year_filter=args.year,
        force=args.force,
    )